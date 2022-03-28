"""
Author: Roomey Rahman
mail: roomeyrahman@gmail.com
"""

import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import json
from django.core.cache import cache

from django.http import HttpResponse


class Report:
    border = Border(left=Side(border_style=None,
                              color='FF000000'),
                    right=Side(border_style=None,
                               color='FF000000'),
                    top=Side(border_style=None,
                             color='FF000000'),
                    bottom=Side(border_style=None,
                                color='FF000000'),
                    diagonal=Side(border_style=None,
                                  color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style=None,
                                 color='FF000000'),
                    vertical=Side(border_style=None,
                                  color='FF000000'),
                    horizontal=Side(border_style=None,
                                    color='FF000000')
                    )

    def __init__(self, df='', jsonObject='', header='', max_row=0, font='', fill='', border='', alignment='',
                 number_format='General', protection=''):
        cache.clear()
        if type(jsonObject) == str and jsonObject != '' and type(df) != str:
            raise Exception("Multiple object send. You must send either dataframe of a JSON object")

        elif type(df) != str:
            if isinstance(df, pd.DataFrame) == True:
                self.df = df
            else:
                raise Exception("Type mismatch. Requires queryset must be of type of 'DataFrame'")
        elif type(jsonObject) == str and jsonObject != '':
            try:
                jsonObject = json.loads(jsonObject)
                self.df = pd.io.json.json_normalize(jsonObject)
            except:
                raise Exception(
                    "Type mismatch. Requires queryset must be of type of either 'DataFrame' or 'Json Object'")

        if type(header) != str:
            if (type(header) == list):
                self.header = header
            else:
                raise Exception("Type mismatch. header must be a list.")
        if max_row > 0:
            self.max_row = max_row - 1
        else:
            self.max_row=0

        
        if type(font) == str and font == '':
            self.font = Font(name='Calibri',
                             size=11,
                             bold=False,
                             italic=False,
                             vertAlign=None,
                             underline='none',
                             strike=False,
                             color='FF000000')
        elif type(font) == dict:
            self.font = Font(
                name=font.get('name') if 'name' in font else 'Calibri',
                size=font.get('size') if 'size' in font else 11,
                bold=font.get('bold') if 'bold' in font else False,
                italic=font.get('italic') if 'italic' in font else False,
                vertAlign=font.get('vertAlign') if 'vertAlign' in font else None,
                underline=font.get('underline') if 'underline' in font else 'none',
                strike=font.get('strike') if 'strike' in font else False,
                color=font.get('color') if 'color' in font else 'FF000000'
            )
        else:
            raise Exception("Type mismatch. Font must be a type of dictionary")

        if type(fill) == str and fill == '':
            self.fill = PatternFill(fill_type=None,
                                    start_color='FFFFFFFF',
                                    end_color='FF000000')
        elif type(fill) == dict:
            self.fill = PatternFill(
                fill_type=fill.get('fill_type') if 'fill_type' in fill else None,
                start_color=fill.get('start_color') if 'start_color' in fill else 11,
                end_color=fill.get('end_color') if 'end_color' in fill else False,
            )
        else:
            raise Exception("Type mismatch. Fill must be a type of dictionary")

        if type(alignment) == str and alignment == '':
            self.alignment = Alignment(horizontal='general',
                                       vertical='bottom',
                                       text_rotation=0,
                                       wrap_text=False,
                                       shrink_to_fit=False,
                                       indent=0)
        elif type(alignment) == dict:
            self.alignment = Alignment(horizontal=alignment.get('horizontal') if 'horizontal' in alignment else 'general',
                                       vertical=alignment.get('vertical') if 'vertical' in alignment else 'bottom',
                                       text_rotation=alignment.get(
                                           'text_rotation') if 'text_rotation' in alignment else 0,
                                       wrap_text=alignment.get('wrap_text') if 'wrap_text' in alignment else False,
                                       shrink_to_fit=alignment.get(
                                           'shrink_to_fit') if 'shrink_to_fit' in alignment else False,
                                       indent=alignment.get('indent') if 'indent' in alignment else 0)
        else:
            raise Exception("Type mismatch. alignment must be a type of dictionary")

        if type(protection) == str and protection == '':
            self.protection = Protection(locked=True,
                                         hidden=False)
        elif type(protection) == dict:
            self.protection = Protection(locked=protection.get('locked') if 'locked' in protection else True,
                                         hidden=protection.get('hidden') if 'hidden' in protection else False
                                         )
        else:
            raise Exception("Type mismatch. protection must be a type of dictionary")

    def exportToExcel(self):
        response = HttpResponse(content_type='application/ms-excel')

        # decide file name
        response['Content-Disposition'] = 'attachment; filename="excel-report.xlsx"'

        # creating workbook
        wb = Workbook()

        # grab the active worksheet
        ws = wb.active
        max_row = -1

        for i in self.header:
            try:
                cell = i['column']
                title = i['title']

                font = i.get('font', self.font)

                font_size = font.get('font_size') if 'font-size' in font else self.font.size
                font_family = font.get('font_family') if 'font_family' in font else self.font.name
                bold = font.get('bold') if 'bold' in font else self.font.bold
                italic = font.get('italic') if 'italic' in font else self.font.italic
                underline = font.get('underline') if 'underline' in font else self.font.underline
                color = font.get('color') if 'color' in font else self.font.color

                alignment = i.get('alignment', self.alignment)

                horizontalAlign = alignment.get(
                    'horizontal') if 'horizontal' in alignment else self.alignment.horizontal
                verticalAlign = alignment.get('vertical') if 'vertical' in alignment else self.alignment.vertical

                ws.merge_cells(cell)
                cell_splt = cell.split(':')
                cell_l = cell_splt[0]
                cell_r = cell_splt[1]

                if self.max_row<=0:
                    max_row = max(max_row, int(re.match(r"([a-z]+)([0-9]+)", cell_l, re.I).groups()[1]))
                    max_row = max(max_row, int(re.match(r"([a-z]+)([0-9]+)", cell_r, re.I).groups()[1]))


                ws[cell_l] = title

                ws[cell_l] = title
                ws[cell_l].font = Font(name=font_family,
                                       size=font_size,
                                       bold=bold,
                                       italic=italic,
                                       underline=underline,
                                       color=color
                                       )
                ws[cell_l].alignment = Alignment(horizontal=horizontalAlign,
                                                 vertical=verticalAlign
                                                 )

                # ws.column_dimensions['A'].width = 10000000
            except Exception as e:
                import logging
                logger = logging.getLogger('ftpuploader')
                logger.error('Failed: ' + str(e))

        rows = dataframe_to_rows(self.df, header=False, index=False)
        
        if self.max_row>0:
            max_row=self.max_row
        skip =max_row
        for r_idx, row in enumerate(rows, 1):
            if row[0] == 'None':
                skip+=1
                continue
            else:
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx + skip , column=c_idx, value=value)
                    

        wb.save(response)

        return response
